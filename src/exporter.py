"""
exporter.py — .xlsx / .csv export with the dated filename convention.
Ekspor data ke Excel/CSV. Nama file: SPEEDHOME_<Area>_<YYYYMMDD>.xlsx

All functions return in-memory bytes so they plug straight into Streamlit's
st.download_button without writing to disk.
"""

from __future__ import annotations

import io
import re
from datetime import date

import pandas as pd

# (raw_key, Display Header) — order matters for the exported table.
LISTING_COLUMNS = [
    ("title", "Listing Title"),
    ("property_name", "Property"),
    ("area", "Area"),
    ("unit_type", "Type"),
    ("bedrooms", "Bedrooms"),
    ("bathrooms", "Bathrooms"),
    ("price_month", "Price / Month (RM)"),
    ("price_year", "Price / Year (RM)"),
    ("price_day", "Price / Day (RM)"),
    ("size_sqft", "Size (sqft)"),
    ("furnishing", "Furnishing"),
    ("rental_type", "Rental Type"),
    ("zero_deposit", "Zero Deposit"),
    ("url", "Listing URL"),
]

SUMMARY_COLUMNS = [
    ("segment", "Unit Type"),
    ("count", "Units Found"),
    ("average", "Average (RM)"),
    ("median", "Median (RM)"),
    ("mode", "Mode (RM)"),
    ("fair_price", "Fair Price (RM)"),
    ("avg_sqft", "Avg Size (sqft)"),
    ("price_psf", "RM / sqft"),
]


def build_filename(area_label: str, ext: str = "xlsx") -> str:
    """SPEEDHOME_Mont_Kiara_20260627.xlsx"""
    area = re.sub(r"[^A-Za-z0-9]+", "_", (area_label or "data").strip()).strip("_")
    stamp = date.today().strftime("%Y%m%d")
    return f"SPEEDHOME_{area}_{stamp}.{ext}"


def listings_dataframe(listings: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(listings)
    for key, _ in LISTING_COLUMNS:
        if key not in df.columns:
            df[key] = None
    df = df[[k for k, _ in LISTING_COLUMNS]]
    df.columns = [h for _, h in LISTING_COLUMNS]
    return df


def summary_dataframe(summary_rows: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(summary_rows)
    for key, _ in SUMMARY_COLUMNS:
        if key not in df.columns:
            df[key] = None
    df = df[[k for k, _ in SUMMARY_COLUMNS]]
    df.columns = [h for _, h in SUMMARY_COLUMNS]
    return df


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")  # BOM => Excel opens UTF-8 cleanly


def to_xlsx_bytes(listings_df: pd.DataFrame, summary_df: pd.DataFrame,
                  meta: dict | None = None) -> bytes:
    """Two-sheet workbook: 'Price Summary' + 'Listings', lightly formatted."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Price Summary", startrow=1)
        listings_df.to_excel(writer, index=False, sheet_name="Listings", startrow=1)
        _format_sheet(writer, "Price Summary", summary_df, title="Price Summary", meta=meta)
        _format_sheet(writer, "Listings", listings_df, title="Unit Listings", meta=meta)
    buffer.seek(0)
    return buffer.getvalue()


def _format_sheet(writer, sheet_name: str, df: pd.DataFrame,
                  title: str = "", meta: dict | None = None) -> None:
    """Bold header row, sensible column widths, a title line, frozen panes."""
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return
    ws = writer.sheets[sheet_name]

    # Title line in A1
    src = (meta or {}).get("area_label", "")
    when = (meta or {}).get("snapshot", date.today().isoformat())
    ws["A1"] = f"{title} — {src}  ·  generated {when}  ·  PriceLens / SPEEDHOME"
    ws["A1"].font = Font(bold=True, size=12, color="FF9A5B3D")  # aRGB (8-hex) for openpyxl

    # Header styling (row 2 because of startrow=1)
    header_fill = PatternFill("solid", fgColor="FFEFEAE1")
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FF23211E")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    # Column widths from content length
    for i, col in enumerate(df.columns, start=1):
        longest = max([len(str(col))] + [len(str(v)) for v in df.iloc[:, i - 1].tolist()] + [8])
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = min(longest + 3, 60)

    ws.freeze_panes = "A3"  # keep header visible while scrolling
